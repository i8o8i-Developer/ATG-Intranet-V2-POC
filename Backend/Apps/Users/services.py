from pathlib import Path
from uuid import uuid4

from django.conf import settings
from decimal import Decimal

from django.db import models
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from Backend.Apps.Users.models import (
    BenchPeriod,
    Department,
    DepartmentMembership,
    EmployeePaymentSnapshot,
    EmployeeProfile,
    InterviewProgress,
    LeaveBalance,
    LeavePolicy,
    LeaveTransaction,
    PayProfile,
    ResignationRequest,
    Skill,
    UserEffortReport,
    UserSkill,
    UserStatusSnapshot,
)
from Backend.EnterpriseCore.services import OutboxService, ServiceResult
from Backend.Apps.MainApp.services import NotificationService


class EmployeeLifecycleService:
    @staticmethod
    def activate_employee(context, employee_id):
        employee = EmployeeProfile.objects.filter(tenant=context.tenant, id=employee_id).first()
        if not employee:
            return ServiceResult.failure({"employee": "Employee Not Found."}, status_code=404)
        employee.status = EmployeeProfile.STATUS_ACTIVE
        employee.updated_by = context.actor
        employee.save(update_fields=["status", "updated_by", "updated_at"])
        OutboxService.publish(context, "EmployeeProfile", employee.id, "EmployeeActivated", {"employeeId": employee.id})
        return ServiceResult.success(employee)

    @staticmethod
    def change_status(context, employee_id, status, reason="", effective_from=None):
        employee = EmployeeProfile.objects.filter(tenant=context.tenant, id=employee_id).first()
        if not employee:
            return ServiceResult.failure({"employee": "Employee Not Found."}, status_code=404)
        effective_from = effective_from or timezone.localdate()
        employee.status = status
        if status == EmployeeProfile.STATUS_EXITED and not employee.exited_on:
            employee.exited_on = effective_from
        employee.updated_by = context.actor
        employee.save(update_fields=["status", "exited_on", "updated_by", "updated_at"])
        snapshot = UserStatusSnapshot.objects.create(
            tenant=context.tenant,
            workspace=context.workspace or employee.workspace,
            employee=employee,
            status=status,
            reason=reason,
            effective_from=effective_from,
            created_by=context.actor,
            updated_by=context.actor,
        )
        if status == EmployeeProfile.STATUS_ON_BENCH:
            bench_period = BenchPeriod.objects.filter(tenant=context.tenant, employee=employee, ended_on__isnull=True).first()
            if not bench_period:
                BenchPeriod.objects.create(
                    tenant=context.tenant,
                    workspace=context.workspace or employee.workspace,
                    employee=employee,
                    reason=reason,
                    created_by=context.actor,
                    updated_by=context.actor,
                )
        else:
            BenchPeriod.objects.filter(tenant=context.tenant, employee=employee, ended_on__isnull=True).update(ended_on=effective_from, updated_by=context.actor)
        OutboxService.publish(context, "EmployeeProfile", employee.id, "EmployeeStatusChanged", {"employeeId": employee.id, "status": status, "snapshotId": snapshot.id})
        return ServiceResult.success(employee)

    @staticmethod
    def transfer_department(context, employee_id, department_id, sub_department_id=None, started_on=None, end_existing=True):
        employee = EmployeeProfile.objects.filter(tenant=context.tenant, id=employee_id).first()
        if not employee:
            return ServiceResult.failure({"employee": "Employee Not Found."}, status_code=404)
        department = Department.objects.filter(tenant=context.tenant, id=department_id).first()
        if not department:
            return ServiceResult.failure({"department": "Department Not Found."}, status_code=404)
        if end_existing:
            DepartmentMembership.objects.filter(tenant=context.tenant, employee=employee, status=DepartmentMembership.STATUS_ACTIVE).update(
                status=DepartmentMembership.STATUS_ENDED,
                ended_on=started_on or timezone.localdate(),
                updated_by=context.actor,
            )
        membership = DepartmentMembership.objects.create(
            tenant=context.tenant,
            workspace=context.workspace or employee.workspace,
            employee=employee,
            department=department,
            sub_department_id=sub_department_id,
            started_on=started_on or timezone.localdate(),
            created_by=context.actor,
            updated_by=context.actor,
        )
        employee.department = department
        employee.updated_by = context.actor
        employee.save(update_fields=["department", "updated_by", "updated_at"])
        EmployeeLifecycleService.assign_department_skills(context, employee)
        OutboxService.publish(context, "EmployeeProfile", employee.id, "EmployeeDepartmentChanged", {"employeeId": employee.id, "departmentId": department.id})
        return ServiceResult.success(membership, status_code=201)

    @staticmethod
    def assign_skill(context, employee, skill, proficiency=1, rating=0, assigned_from_department=False):
        link, _created = UserSkill.objects.update_or_create(
            tenant=context.tenant,
            workspace=context.workspace or employee.workspace,
            employee=employee,
            skill=skill,
            defaults={"proficiency": proficiency, "rating": rating, "assigned_from_department": assigned_from_department, "updated_by": context.actor},
        )
        
        # Create Notification For Employee
        if employee.user and not assigned_from_department:
            proficiency_label = {1: "Basic", 2: "Intermediate", 3: "Advanced"}.get(proficiency, "Basic")
            NotificationService.notify(
                context,
                recipient=employee.user,
                title=f"Skill Updated: {skill.name}",
                message=f"Your Skill Level For {skill.name} Has Been Updated To {proficiency_label}.",
                category="hrms",
                resource_type="skill",
                resource_id=str(link.id),
                metadata={"employee_id": str(employee.id), "skill_id": str(skill.id), "proficiency": proficiency}
            )
        
        return ServiceResult.success(link)

    @staticmethod
    def assign_department_skills(context, employee):
        department_ids = list(
            DepartmentMembership.objects.filter(tenant=context.tenant, employee=employee, status=DepartmentMembership.STATUS_ACTIVE).values_list("department_id", flat=True)
        )
        if employee.department_id and employee.department_id not in department_ids:
            department_ids.append(employee.department_id)
        skills = Skill.objects.filter(tenant=context.tenant, is_active=True).filter(models.Q(department_id__in=department_ids) | models.Q(department__isnull=True))
        assigned = []
        for skill in skills:
            result = EmployeeLifecycleService.assign_skill(context, employee, skill, proficiency=1, assigned_from_department=bool(skill.department_id))
            if result.ok:
                assigned.append(result.data.id)
        return ServiceResult.success({"employeeId": employee.id, "assignedSkillIds": assigned, "count": len(assigned)})

    @staticmethod
    def assign_all_department_skills(context):
        employees = EmployeeProfile.objects.filter(tenant=context.tenant, status=EmployeeProfile.STATUS_ACTIVE, is_active=True)
        rows = []
        for employee in employees:
            rows.append(EmployeeLifecycleService.assign_department_skills(context, employee).data)
        return ServiceResult.success({"count": len(rows), "employees": rows})

    @staticmethod
    def complete_onboarding(context, employee_id):
        employee = EmployeeProfile.objects.filter(tenant=context.tenant, id=employee_id).first()
        if not employee:
            return ServiceResult.failure({"employee": "Employee Not Found."}, status_code=404)
        employee.onboarding_completed = True
        employee.updated_by = context.actor
        employee.save(update_fields=["onboarding_completed", "updated_by", "updated_at"])
        return ServiceResult.success(employee)

    @staticmethod
    def save_timezone(context, employee_id, timezone_name):
        employee = EmployeeProfile.objects.filter(tenant=context.tenant, id=employee_id).first()
        if not employee:
            return ServiceResult.failure({"employee": "Employee Not Found."}, status_code=404)
        employee.timezone_name = timezone_name
        employee.updated_by = context.actor
        employee.save(update_fields=["timezone_name", "updated_by", "updated_at"])
        return ServiceResult.success(employee)


class LeaveWalletService:
    @staticmethod
    def get_or_create_policy(context):
        policy, _created = LeavePolicy.objects.get_or_create(
            tenant=context.tenant,
            workspace=context.workspace,
            name="Default",
            defaults={"leaves_per_month": 2.25, "created_by": context.actor, "updated_by": context.actor},
        )
        return policy

    @staticmethod
    def accrue_for_employee(context, employee, policy=None, amount=None, reason="Monthly Leave Accrual"):
        policy = policy or LeaveWalletService.get_or_create_policy(context)
        amount = Decimal(str(amount)) if amount is not None else policy.leaves_per_month
        balance, _created = LeaveBalance.objects.get_or_create(
            tenant=context.tenant,
            workspace=context.workspace or employee.workspace,
            employee=employee,
            policy=policy,
            defaults={"created_by": context.actor, "updated_by": context.actor},
        )
        balance.available += amount
        balance.accrued += amount
        balance.save(update_fields=["available", "accrued", "updated_at"])
        employee.leaves_wallet = balance.available
        employee.leaves_per_month = policy.leaves_per_month
        employee.save(update_fields=["leaves_wallet", "leaves_per_month", "updated_at"])
        transaction = LeaveTransaction.objects.create(
            tenant=context.tenant,
            workspace=context.workspace or employee.workspace,
            balance=balance,
            transaction_type="Accrual",
            amount=amount,
            reason=reason,
            created_by=context.actor,
            updated_by=context.actor,
        )
        return ServiceResult.success(transaction, status_code=201)

    @staticmethod
    def update_all_wallets(context, amount=None):
        policy = LeaveWalletService.get_or_create_policy(context)
        transactions = []
        for employee in EmployeeProfile.objects.filter(tenant=context.tenant, status=EmployeeProfile.STATUS_ACTIVE, is_active=True):
            result = LeaveWalletService.accrue_for_employee(context, employee, policy=policy, amount=amount)
            if result.ok:
                transactions.append(result.data.id)
        return ServiceResult.success({"count": len(transactions), "transactionIds": transactions})


class UserWorkflowService:
    @staticmethod
    def submit_effort_report(context, employee_id, report_month, report_year, effort_percent, project_reference="", metadata=None):
        employee = EmployeeProfile.objects.filter(tenant=context.tenant, id=employee_id).first()
        if not employee:
            return ServiceResult.failure({"employee": "Employee Not Found."}, status_code=404)
        report = UserEffortReport.objects.create(
            tenant=context.tenant,
            workspace=context.workspace or employee.workspace,
            employee=employee,
            report_month=report_month,
            report_year=report_year,
            effort_percent=effort_percent,
            project_reference=project_reference,
            metadata=metadata or {},
            created_by=context.actor,
            updated_by=context.actor,
        )
        OutboxService.publish(context, "UserEffortReport", report.id, "EffortReportSubmitted", {"employeeId": employee.id})
        return ServiceResult.success(report, status_code=201)

    @staticmethod
    def create_effort_report_reminders(context, report_month=None, report_year=None):
        now = timezone.localdate()
        report_month = report_month or now.month
        report_year = report_year or now.year
        employees = EmployeeProfile.objects.filter(tenant=context.tenant, status=EmployeeProfile.STATUS_ACTIVE, is_active=True)
        fixed_pay_employee_ids = PayProfile.objects.filter(tenant=context.tenant, pay_type__iexact="Fixed").values_list("employee_id", flat=True)
        employees = employees.filter(id__in=fixed_pay_employee_ids)
        event_ids = []
        for employee in employees:
            exists = UserEffortReport.objects.filter(tenant=context.tenant, employee=employee, report_month=report_month, report_year=report_year).exists()
            if exists:
                continue
            event = OutboxService.publish(
                context,
                "EmployeeProfile",
                employee.id,
                "EffortReportReminderCreated",
                {"employeeId": employee.id, "month": report_month, "year": report_year},
            )
            event_ids.append(event.id)
        return ServiceResult.success({"count": len(event_ids), "eventIds": event_ids})

    @staticmethod
    def create_daily_activity_reminders(context, reminder_type="EOD"):
        event_ids = []
        for employee in EmployeeProfile.objects.filter(tenant=context.tenant, status=EmployeeProfile.STATUS_ACTIVE, is_active=True):
            event = OutboxService.publish(
                context,
                "EmployeeProfile",
                employee.id,
                "DailyActivityReminderCreated",
                {"employeeId": employee.id, "reminderType": reminder_type},
            )
            event_ids.append(event.id)
        return ServiceResult.success({"count": len(event_ids), "eventIds": event_ids})

    @staticmethod
    def submit_resignation(context, employee_id, reason, last_working_day=None):
        employee = EmployeeProfile.objects.filter(tenant=context.tenant, id=employee_id).first()
        if not employee:
            return ServiceResult.failure({"employee": "Employee Not Found."}, status_code=404)
        resignation = ResignationRequest.objects.create(
            tenant=context.tenant,
            workspace=context.workspace or employee.workspace,
            employee=employee,
            reason=reason,
            last_working_day=last_working_day,
            created_by=context.actor,
            updated_by=context.actor,
        )
        OutboxService.publish(context, "ResignationRequest", resignation.id, "ResignationSubmitted", {"employeeId": employee.id})
        return ServiceResult.success(resignation, status_code=201)

    @staticmethod
    def approve_resignation(context, resignation_id, last_working_day=None):
        resignation = ResignationRequest.objects.filter(tenant=context.tenant, id=resignation_id).select_related("employee").first()
        if not resignation:
            return ServiceResult.failure({"resignation": "Resignation Request Not Found."}, status_code=404)
        resignation.status = "Approved"
        resignation.approved_by = context.actor
        resignation.approved_at = timezone.now()
        if last_working_day:
            resignation.last_working_day = last_working_day
        resignation.save(update_fields=["status", "approved_by", "approved_at", "last_working_day", "updated_at"])
        EmployeeLifecycleService.change_status(context, resignation.employee_id, EmployeeProfile.STATUS_EXITED, reason="Resignation approved", effective_from=resignation.last_working_day)
        OutboxService.publish(context, "ResignationRequest", resignation.id, "ResignationApproved", {"employeeId": resignation.employee_id})
        return ServiceResult.success(resignation)


class InterviewSyncService:
    @staticmethod
    def sync_interns(context, employee_id=None, dry_run=True, send_links=False, client=None, mode="sync"):
        from Backend.Apps.Users.interviewgod import InterviewGodService

        return InterviewGodService.run_scheduler(context, employee_id=employee_id, mode=mode, dry_run=dry_run, send_links=send_links, client=client)


class PaymentSyncService:
    @staticmethod
    def request_payment_status_sync(context):
        event_ids = []
        payments = EmployeePaymentSnapshot.objects.filter(tenant=context.tenant).exclude(payout_id="").exclude(payment_status__iexact="processed")
        for payment in payments:
            event = OutboxService.publish(context, "EmployeePaymentSnapshot", payment.id, "PaymentStatusSyncRequested", {"payoutId": payment.payout_id})
            event_ids.append(event.id)
        return ServiceResult.success({"count": len(event_ids), "eventIds": event_ids})


class PayrollExportService:
    EXPORT_FOLDER = "Temp_Payroll"
    HEADERS = [
        "Name",
        "Employee Code",
        "Department",
        "Date Of Joining",
        "Bank Account Number",
        "Account IFSC",
        "UPI Id",
        "Pay Type",
        "Base Pay",
        "Bonus",
        "Deduction",
        "Bounty",
        "Net Pay",
        "Leaves",
        "Effort Percent",
        "UTR",
        "Payment Status",
        "Finance Status",
        "Manager Status",
        "Additional Remarks",
    ]

    @staticmethod
    def generate_excel(context, report_month=None, report_year=None, pay_type=""):
        if not context.tenant:
            return ServiceResult.failure({"tenant": "Tenant Context Is Required."}, status_code=400)

        report_month = int(report_month) if report_month else None
        report_year = int(report_year) if report_year else None
        rows = PayrollExportService._build_rows(context, report_month=report_month, report_year=report_year, pay_type=pay_type)

        export_dir = Path(settings.MEDIA_ROOT) / PayrollExportService.EXPORT_FOLDER
        export_dir.mkdir(parents=True, exist_ok=True)
        filename = PayrollExportService._build_filename(context, report_month=report_month, report_year=report_year)
        file_path = export_dir / filename

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Payroll Data"
        worksheet.append(PayrollExportService.HEADERS)
        PayrollExportService._style_header(worksheet)
        for row in rows:
            worksheet.append([row.get(header, "") for header in PayrollExportService.HEADERS])
        PayrollExportService._autosize(worksheet)
        workbook.save(file_path)

        return ServiceResult.success(
            {
                "filename": filename,
                "rowCount": len(rows),
                "generatedAt": timezone.now().isoformat(),
                "filters": {"month": report_month, "year": report_year, "pay_type": pay_type or ""},
            }
        )

    @staticmethod
    def resolve_export_path(context, filename):
        if not context.tenant:
            return None
        safe_name = Path(filename).name
        expected_prefix = f"tenant-{context.tenant.id}-payroll-"
        if not safe_name.startswith(expected_prefix):
            return None
        export_path = Path(settings.MEDIA_ROOT) / PayrollExportService.EXPORT_FOLDER / safe_name
        return export_path if export_path.exists() else None

    @staticmethod
    def _build_rows(context, report_month=None, report_year=None, pay_type=""):
        employees = EmployeeProfile.objects.filter(tenant=context.tenant, status=EmployeeProfile.STATUS_ACTIVE, is_active=True).select_related(
            "user",
            "department",
        ).prefetch_related(
            "pay_profiles",
            "user_bank_accounts",
            "payment_snapshots",
            "leave_balances",
            "effort_reports",
        )
        if context.workspace:
            employees = employees.filter(workspace=context.workspace)

        period_end = PayrollExportService._period_end(report_month, report_year)
        if period_end:
            employees = employees.filter(models.Q(joined_on__isnull=True) | models.Q(joined_on__lte=period_end))

        rows = []
        for employee in employees:
            pay_profile = PayrollExportService._pick_pay_profile(employee, period_end=period_end)
            if pay_type and (not pay_profile or pay_profile.pay_type.lower() != pay_type.lower()):
                continue
            payment_snapshot = PayrollExportService._pick_payment_snapshot(employee, report_month=report_month, report_year=report_year)
            bank_account = PayrollExportService._pick_first(employee.user_bank_accounts.all())
            leave_balance = PayrollExportService._pick_first(employee.leave_balances.all())
            effort_report = PayrollExportService._pick_effort_report(employee, report_month=report_month, report_year=report_year)

            base_pay = pay_profile.base_pay if pay_profile else Decimal("0")
            bonus = payment_snapshot.bonus if payment_snapshot else Decimal("0")
            deduction = payment_snapshot.deduction if payment_snapshot else Decimal("0")
            bounty = payment_snapshot.bounty if payment_snapshot else Decimal("0")
            normal_pay = payment_snapshot.normal_pay if payment_snapshot else base_pay
            net_pay = normal_pay + bonus + bounty - deduction

            rows.append(
                {
                    "Name": employee.display_name or employee.user.get_full_name() or employee.user.username,
                    "Employee Code": employee.employee_code,
                    "Department": employee.department.name if employee.department_id else "",
                    "Date Of Joining": employee.joined_on.isoformat() if employee.joined_on else "",
                    "Bank Account Number": bank_account.masked_account_number if bank_account else "",
                    "Account IFSC": bank_account.ifsc_code if bank_account else "",
                    "UPI Id": bank_account.upi_id if bank_account else "",
                    "Pay Type": pay_profile.pay_type if pay_profile else "",
                    "Base Pay": str(base_pay),
                    "Bonus": str(bonus),
                    "Deduction": str(deduction),
                    "Bounty": str(bounty),
                    "Net Pay": str(net_pay),
                    "Leaves": str(leave_balance.available if leave_balance else employee.leaves_wallet),
                    "Effort Percent": str(effort_report.effort_percent if effort_report else Decimal("0")),
                    "UTR": payment_snapshot.utr_number if payment_snapshot else "",
                    "Payment Status": payment_snapshot.payment_status if payment_snapshot else "",
                    "Finance Status": payment_snapshot.finance_status if payment_snapshot else "",
                    "Manager Status": payment_snapshot.manager_status if payment_snapshot else "",
                    "Additional Remarks": payment_snapshot.notes if payment_snapshot else "",
                }
            )
        return rows

    @staticmethod
    def _pick_first(items):
        return next(iter(items), None)

    @staticmethod
    def _pick_pay_profile(employee, period_end=None):
        profiles = sorted(employee.pay_profiles.all(), key=lambda item: item.effective_at, reverse=True)
        if period_end:
            for profile in profiles:
                if profile.effective_at.date() <= period_end:
                    return profile
        return PayrollExportService._pick_first(profiles)

    @staticmethod
    def _pick_payment_snapshot(employee, report_month=None, report_year=None):
        snapshots = sorted(employee.payment_snapshots.all(), key=lambda item: (item.year, item.month, item.id), reverse=True)
        if report_month and report_year:
            for snapshot in snapshots:
                if snapshot.month == report_month and snapshot.year == report_year:
                    return snapshot
            return None
        return PayrollExportService._pick_first(snapshots)

    @staticmethod
    def _pick_effort_report(employee, report_month=None, report_year=None):
        reports = sorted(employee.effort_reports.all(), key=lambda item: (item.report_year, item.report_month, item.id), reverse=True)
        if report_month and report_year:
            for report in reports:
                if report.report_month == report_month and report.report_year == report_year:
                    return report
            return None
        return PayrollExportService._pick_first(reports)

    @staticmethod
    def _period_end(report_month, report_year):
        if not report_month or not report_year:
            return None
        if report_month == 12:
            return timezone.datetime(report_year, 12, 31).date()
        return (timezone.datetime(report_year, report_month + 1, 1) - timezone.timedelta(days=1)).date()

    @staticmethod
    def _build_filename(context, report_month=None, report_year=None):
        month_label = str(report_month or "latest")
        year_label = str(report_year or "latest")
        return f"tenant-{context.tenant.id}-payroll-{year_label}-{month_label}-{uuid4().hex[:8]}.xlsx"

    @staticmethod
    def _style_header(worksheet):
        fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        font = Font(bold=True, color="FFFFFF")
        alignment = Alignment(horizontal="center", vertical="center")
        for cell in worksheet[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = alignment

    @staticmethod
    def _autosize(worksheet):
        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 36)


class HRMSDashboardService:
    @staticmethod
    def summarize(context):
        employees = EmployeeProfile.objects.filter(tenant=context.tenant, is_active=True)
        by_status = list(employees.values("status").annotate(count=models.Count("id")).order_by("status"))
        by_department = list(employees.values("department__name").annotate(count=models.Count("id")).order_by("department__name"))
        open_resignations = ResignationRequest.objects.filter(tenant=context.tenant, status="Submitted").count()
        open_bench_periods = BenchPeriod.objects.filter(tenant=context.tenant, ended_on__isnull=True).count()
        return ServiceResult.success(
            {
                "headcount": employees.count(),
                "byStatus": by_status,
                "byDepartment": by_department,
                "openResignations": open_resignations,
                "openBenchPeriods": open_bench_periods,
            }
        )
